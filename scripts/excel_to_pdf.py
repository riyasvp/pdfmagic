#!/usr/bin/env python3
"""
Convert Excel file to PDF.
Usage: python excel_to_pdf.py <input_excel>
Output: JSON with result
"""

import sys
import os
import json
from datetime import datetime

try:
    from openpyxl import load_workbook
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
    from reportlab.lib import colors
    from reportlab.lib.units import inch
except ImportError as e:
    print(json.dumps({"success": False, "error": f"Missing dependency: {str(e)}"}))
    sys.exit(1)

DOWNLOAD_DIR = "/home/z/my-project/download"

def excel_to_pdf(input_path):
    """Convert Excel file to PDF."""
    if not os.path.exists(input_path):
        return {"success": False, "error": f"File not found: {input_path}"}
    
    try:
        # Load Excel workbook
        wb = load_workbook(input_path)
        
        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"excel_to_pdf_{timestamp}.pdf"
        output_path = os.path.join(DOWNLOAD_DIR, output_filename)
        
        # Ensure download directory exists
        os.makedirs(DOWNLOAD_DIR, exist_ok=True)
        
        # Create PDF document
        doc = SimpleDocTemplate(output_path, pagesize=landscape(A4))
        story = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Add sheet name as header
            from reportlab.platypus import Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            styles = getSampleStyleSheet()
            story.append(Paragraph(f"<b>Sheet: {sheet_name}</b>", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Get data from sheet
            data = []
            for row in sheet.iter_rows(values_only=True):
                # Convert None to empty string and handle non-string values
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(cell for cell in row_data):  # Skip empty rows
                    data.append(row_data)
            
            if data:
                # Create table
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                story.append(table)
            
            story.append(Spacer(1, 30))
        
        doc.build(story)
        
        return {"success": True, "output": output_path}
    
    except Exception as e:
        return {"success": False, "error": str(e)}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "Input Excel file required"}))
        sys.exit(1)
    
    input_path = sys.argv[1]
    result = excel_to_pdf(input_path)
    print(json.dumps(result))
