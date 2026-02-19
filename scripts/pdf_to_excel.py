#!/usr/bin/env python3
"""
Convert PDF to Excel.
Usage: python pdf_to_excel.py <input_pdf>
Output: JSON with result
"""

import sys
import os
import json
from datetime import datetime

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError as e:
    print(json.dumps({"success": False, "error": f"Missing dependency: {str(e)}"}))
    sys.exit(1)

DOWNLOAD_DIR = "/home/z/my-project/download"

def pdf_to_excel(input_path):
    """Convert PDF tables to Excel."""
    if not os.path.exists(input_path):
        return {"success": False, "error": f"File not found: {input_path}"}
    
    try:
        # Create workbook
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        tables_found = 0
        
        with pdfplumber.open(input_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()
                
                if tables:
                    for table_num, table in enumerate(tables, 1):
                        tables_found += 1
                        # Create sheet for each table
                        sheet_name = f"Page{page_num}_Table{table_num}"[:31]  # Excel sheet name limit
                        ws = wb.create_sheet(title=sheet_name)
                        
                        # Style definitions
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for row_idx, row in enumerate(table, 1):
                            for col_idx, cell in enumerate(row, 1):
                                cell_value = str(cell) if cell is not None else ""
                                ws.cell(row=row_idx, column=col_idx, value=cell_value)
                                
                                # Apply styles
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                
                                if row_idx == 1:
                                    cell.font = header_font
                                    cell.fill = header_fill
        
        if tables_found == 0:
            # If no tables found, extract text
            ws = wb.create_sheet(title="Text Content")
            with pdfplumber.open(input_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    text = page.extract_text()
                    if text:
                        ws.append([f"--- Page {page_num} ---"])
                        for line in text.split('\n'):
                            ws.append([line])
                        ws.append([])
        
        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"pdf_to_excel_{timestamp}.xlsx"
        output_path = os.path.join(DOWNLOAD_DIR, output_filename)
        
        # Ensure download directory exists
        os.makedirs(DOWNLOAD_DIR, exist_ok=True)
        
        # Save workbook
        wb.save(output_path)
        
        return {
            "success": True,
            "output": output_path,
            "tables_extracted": tables_found
        }
    
    except Exception as e:
        return {"success": False, "error": str(e)}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "Input PDF file required"}))
        sys.exit(1)
    
    input_path = sys.argv[1]
    result = pdf_to_excel(input_path)
    print(json.dumps(result))
